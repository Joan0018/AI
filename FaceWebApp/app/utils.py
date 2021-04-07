import time

import numpy as np
# import pandas as pd
# import matplotlib.pyplot as plt
import sklearn
import os
import pickle
import cv2
import face_recognition
import pathlib
import xlsxwriter
from openpyxl import load_workbook
import HandTrackingModule as htm

# Export image of object
OBJECT_IMAGE_EXPORT_PATH = 'static/object_detected/image/'

# Model for face Detection
cascPathface = os.path.dirname(cv2.__file__) + "/data/haarcascade_frontalface_alt2.xml"
haar = cv2.CascadeClassifier(cascPathface)

# pickle files
mean = pickle.load(open('./model/mean_preprocess.pickle', 'rb'))
model_svm = pickle.load(open('./model/model_svm.pickle', 'rb'))
model_pca = pickle.load(open('./model/pca_50.pickle', 'rb'))

print('Model loaded sucessfully')

# settins
gender_pre = ['Male', 'Female']
font = cv2.FONT_HERSHEY_SIMPLEX


# Gender Classification (Image)
def pipeline_model(path, filename, color='bgr'):
    name = faceRecognitionImage(path)

    # step-1: read image in cv2
    img = cv2.imread(path)

    # step-2: convert into gray scale
    if color == 'bgr':
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    else:
        gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)

    # step-3: crop the face (using haar cascade classifier)
    faces = haar.detectMultiScale(gray, 1.5, 3)
    for x, y, w, h in faces:
        cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)  # drawing rectangle
        roi = gray[y:y + h, x:x + w]  # crop image

        # step - 4: normalization (0-1)
        roi = roi / 255.0

        # step-5: resize images (100,100)
        if roi.shape[1] > 100:
            roi_resize = cv2.resize(roi, (100, 100), cv2.INTER_AREA)
        else:
            roi_resize = cv2.resize(roi, (100, 100), cv2.INTER_CUBIC)

        # step-6: Flattening (1x10000)
        roi_reshape = roi_resize.reshape(1, 10000)  # 1,-1

        # step-7: subtract with mean
        roi_mean = roi_reshape - mean

        # step -8: get eigen image
        eigen_image = model_pca.transform(roi_mean)

        # step -9: pass to ml model (svm)
        results = model_svm.predict_proba(eigen_image)[0]

        # step -10:
        predict = results.argmax()  # 0 or 1
        score = results[predict]

        # Get the height and width for image for put text
        imageWidth = img.shape[1]
        imageHeight = img.shape[0]
        # step -11:
        text = " %s : %0.2f" % (gender_pre[predict], score)
        txtName = "%s" % (name)
        scale = 1  # this value can be from 0 to 1 (0,1] to change the size of the text relative to the image
        fontScale = min(imageWidth, imageHeight) / (25 / scale)
        cv2.putText(img, text, (x, y), font, 1, (255, 255, 0), 2)
        cv2.putText(img, txtName, (x, y + w), font, 1, (255, 255, 0), 2, cv2.LINE_AA)

    cv2.imwrite('./static/predict/{}'.format(filename), img)


# Face Recognition
def faceRecognitionImage(path):
    # #loading the image to detect
    global name_of_person
    original_image = cv2.imread(path)

    # #load the sample images and get the 128 face embeddings from them
    # Face #1
    Benedict_Cumberbatch_image = face_recognition.load_image_file('images/New Sample/Benedict_Cumberbatch.jpg')
    Benedict_Cumberbatch_face_encodings = face_recognition.face_encodings(Benedict_Cumberbatch_image)[0]

    # Face #2
    Beyonce_image = face_recognition.load_image_file('images/New Sample/Beyonce.jpg')
    Beyonce_face_encodings = face_recognition.face_encodings(Beyonce_image)[0]

    # Face #3
    lim_image = face_recognition.load_image_file('images/Lim.jpeg')
    lim_face_encodings = face_recognition.face_encodings(lim_image)[0]

    # Face #4
    Brie_Larson_image = face_recognition.load_image_file('images/New Sample/Brie_Larson.jpg')
    Brie_Larson_encodings = face_recognition.face_encodings(Brie_Larson_image)[0]

    # Face #5
    Calum_Scott_image = face_recognition.load_image_file('images/New Sample/Calum  Scott.jpeg')
    Calum_Scott_encodings = face_recognition.face_encodings(Calum_Scott_image)[0]

    # Face #6
    Chris_Hemsworth_image = face_recognition.load_image_file('images/New Sample/Chris Hemsworth.jpg')
    Chris_Hemsworth_encodings = face_recognition.face_encodings(Chris_Hemsworth_image)[0]

    # Face #7
    Chris_Hemsworth_image = face_recognition.load_image_file('images/New Sample/Chris Pratt.jpg')
    Chris_Hemsworth_encodings = face_recognition.face_encodings(Chris_Hemsworth_image)[0]

    # Face #8
    Ed_Sheeran_image = face_recognition.load_image_file('images/New Sample/Ed Sheeran.png')
    Ed_Sheeran_encodings = face_recognition.face_encodings(Ed_Sheeran_image)[0]

    # Face #9
    Evangeline_Lilly_image = face_recognition.load_image_file('images/New Sample/Evangeline_Lilly.jpg')
    Evangeline_Lilly_encodings = face_recognition.face_encodings(Evangeline_Lilly_image)[0]

    # Face #10
    Hailey_Rhode_Bieber_image = face_recognition.load_image_file('images/New Sample/Hailey Rhode Bieber.jpg')
    Hailey_Rhode_Bieber_encodings = face_recognition.face_encodings(Hailey_Rhode_Bieber_image)[0]

    # Face #11
    justin_bieber_image = face_recognition.load_image_file('images/New Sample/justin bieber.jpg')
    justin_bieber_encodings = face_recognition.face_encodings(justin_bieber_image)[0]

    # Face #12
    lisa_image = face_recognition.load_image_file('images/New Sample/lisa.png')
    lisa_encodings = face_recognition.face_encodings(lisa_image)[0]

    # Face #13
    Mahershala_Ali_image = face_recognition.load_image_file('images/New Sample/Mahershala_Ali.jpg')
    Mahershala_Ali_encodings = face_recognition.face_encodings(Mahershala_Ali_image)[0]

    # Face #14
    Paul_Rudd_image = face_recognition.load_image_file('images/New Sample/Paul_Rudd.jpg')
    Paul_Rudd_encodings = face_recognition.face_encodings(Paul_Rudd_image)[0]

    # Face #14
    Paul_Rudd_image = face_recognition.load_image_file('images/New Sample/Richard_Madden.jpg')
    Paul_Rudd_encodings = face_recognition.face_encodings(Paul_Rudd_image)[0]

    # Face #15
    Ryan_Reynolds_image = face_recognition.load_image_file('images/New Sample/Ryan_Reynolds.jpg')
    Ryan_Reynolds_encodings = face_recognition.face_encodings(Ryan_Reynolds_image)[0]

    # Face #16
    Scarlett_Johansson_image = face_recognition.load_image_file('images/New Sample/Scarlett Johansson.jpg')
    Scarlett_Johansson_encodings = face_recognition.face_encodings(Scarlett_Johansson_image)[0]

    # Face #17
    Simu_Liu_image = face_recognition.load_image_file('images/New Sample/Simu_Liu.jpg')
    Simu_Liu_encodings = face_recognition.face_encodings(Simu_Liu_image)[0]

    # Face #18
    taylor_swift_image = face_recognition.load_image_file('images/New Sample/taylor-swift.jpg')
    taylor_swift_encodings = face_recognition.face_encodings(taylor_swift_image)[0]

    # Face #19
    the_rock_image = face_recognition.load_image_file('images/New Sample/the-rock.jpg')
    the_rock_encodings = face_recognition.face_encodings(the_rock_image)[0]

    # Face #20
    Tom_Holland_image = face_recognition.load_image_file('images/New Sample/Tom_Holland.jpg')
    Tom_Holland_encodings = face_recognition.face_encodings(Tom_Holland_image)[0]

    # Face #21
    modi_image = face_recognition.load_image_file('images/modi.jpg')
    modi_encodings = face_recognition.face_encodings(modi_image)[0]

    # #save the encodings and the corresponding labels in seperate arrays in the same order
    known_face_encodings = [Benedict_Cumberbatch_face_encodings, Beyonce_face_encodings, lim_face_encodings,
                            Brie_Larson_encodings, Calum_Scott_encodings, Chris_Hemsworth_encodings,
                            Ed_Sheeran_encodings, Evangeline_Lilly_encodings,
                            Hailey_Rhode_Bieber_encodings, justin_bieber_encodings, lisa_encodings,
                            Mahershala_Ali_encodings,
                            Paul_Rudd_encodings, Ryan_Reynolds_encodings, Scarlett_Johansson_encodings,
                            Simu_Liu_encodings, taylor_swift_encodings,
                            the_rock_encodings, Tom_Holland_encodings, modi_encodings]
    known_face_names = ["Benedict Cumberbatch", "Beyonce", "Lim Kah Yee",
                        "Brie Larson", "Calum Scott", "Chris Hemsworth", "Ed Sheeran", "Evangeline Lilly",
                        "Hailey Rhode Bieber", "Justin Bieber", "Lisa", " Mahershala Ali", "Paul Rudd",
                        "Ryan Reynolds", "Scarlett Johansson", "Simu Liu", "Taylor Swift", "Johnson Dwayne",
                        "Tom Holland", "Narendra Modi"]

    # #load the unknown image to recognize faces in it
    image_to_recognize = face_recognition.load_image_file(path)

    # #detect all faces in the image
    # #arguments are image,no_of_times_to_upsample, model
    all_face_locations = face_recognition.face_locations(image_to_recognize, model='hog')
    # #detect face encodings for all the faces detected
    all_face_encodings = face_recognition.face_encodings(image_to_recognize, all_face_locations)

    # #print the number of faces detected
    print('There are {} no of faces in this image'.format(len(all_face_locations)))

    # #looping through the face locations and the face embeddings
    for current_face_location, current_face_encoding in zip(all_face_locations, all_face_encodings):
        #     #splitting the tuple to get the four position values of current face
        top_pos, right_pos, bottom_pos, left_pos = current_face_location

        #     #find all the matches and get the list of matches
        all_matches = face_recognition.compare_faces(known_face_encodings, current_face_encoding)

        #     #string to hold the label
        name_of_person = 'Unknown face'

        #     #check if the all_matches have at least one item
        #     #if yes, get the index number of face that is located in the first index of all_matches
        #     #get the name corresponding to the index number and save it in name_of_person
        if True in all_matches:
            first_match_index = all_matches.index(True)
            name_of_person = known_face_names[first_match_index]

    return name_of_person


def faceRecognitionVideo(path):
   # capture the video from default camera
    global left_pos, top_pos, right_pos, bottom_pos, name_of_person
    file_video_stream = cv2.VideoCapture(path)

    # load the sample images and get the 128 face embeddings from them
    # Face #1
    Benedict_Cumberbatch_image = face_recognition.load_image_file('images/New Sample/Benedict_Cumberbatch.jpg')
    Benedict_Cumberbatch_face_encodings = face_recognition.face_encodings(Benedict_Cumberbatch_image)[0]

    # Face #2
    Beyonce_image = face_recognition.load_image_file('images/New Sample/Beyonce.jpg')
    Beyonce_face_encodings = face_recognition.face_encodings(Beyonce_image)[0]

    # Face #3
    lim_image = face_recognition.load_image_file('images/Lim.jpeg')
    lim_face_encodings = face_recognition.face_encodings(lim_image)[0]

    # Face #4
    Brie_Larson_image = face_recognition.load_image_file('images/New Sample/Brie_Larson.jpg')
    Brie_Larson_encodings = face_recognition.face_encodings(Brie_Larson_image)[0]

    # Face #5
    Calum_Scott_image = face_recognition.load_image_file('images/New Sample/Calum  Scott.jpeg')
    Calum_Scott_encodings = face_recognition.face_encodings(Calum_Scott_image)[0]

    # Face #6
    Chris_Hemsworth_image = face_recognition.load_image_file('images/New Sample/Chris Hemsworth.jpg')
    Chris_Hemsworth_encodings = face_recognition.face_encodings(Chris_Hemsworth_image)[0]

    # Face #7
    Chris_Hemsworth_image = face_recognition.load_image_file('images/New Sample/Chris Pratt.jpg')
    Chris_Hemsworth_encodings = face_recognition.face_encodings(Chris_Hemsworth_image)[0]

    # Face #8
    Ed_Sheeran_image = face_recognition.load_image_file('images/New Sample/Ed Sheeran.png')
    Ed_Sheeran_encodings = face_recognition.face_encodings(Ed_Sheeran_image)[0]

    # Face #9
    Evangeline_Lilly_image = face_recognition.load_image_file('images/New Sample/Evangeline_Lilly.jpg')
    Evangeline_Lilly_encodings = face_recognition.face_encodings(Evangeline_Lilly_image)[0]

    # Face #10
    Hailey_Rhode_Bieber_image = face_recognition.load_image_file('images/New Sample/Hailey Rhode Bieber.jpg')
    Hailey_Rhode_Bieber_encodings = face_recognition.face_encodings(Hailey_Rhode_Bieber_image)[0]

    # Face #11
    justin_bieber_image = face_recognition.load_image_file('images/New Sample/justin bieber.jpg')
    justin_bieber_encodings = face_recognition.face_encodings(justin_bieber_image)[0]

    # Face #12
    lisa_image = face_recognition.load_image_file('images/New Sample/lisa.png')
    lisa_encodings = face_recognition.face_encodings(lisa_image)[0]

    # Face #13
    Mahershala_Ali_image = face_recognition.load_image_file('images/New Sample/Mahershala_Ali.jpg')
    Mahershala_Ali_encodings = face_recognition.face_encodings(Mahershala_Ali_image)[0]

    # Face #14
    Paul_Rudd_image = face_recognition.load_image_file('images/New Sample/Paul_Rudd.jpg')
    Paul_Rudd_encodings = face_recognition.face_encodings(Paul_Rudd_image)[0]

    # Face #14
    Paul_Rudd_image = face_recognition.load_image_file('images/New Sample/Richard_Madden.jpg')
    Paul_Rudd_encodings = face_recognition.face_encodings(Paul_Rudd_image)[0]

    # Face #15
    Ryan_Reynolds_image = face_recognition.load_image_file('images/New Sample/Ryan_Reynolds.jpg')
    Ryan_Reynolds_encodings = face_recognition.face_encodings(Ryan_Reynolds_image)[0]

    # Face #16
    Scarlett_Johansson_image = face_recognition.load_image_file('images/New Sample/Scarlett Johansson.jpg')
    Scarlett_Johansson_encodings = face_recognition.face_encodings(Scarlett_Johansson_image)[0]

    # Face #17
    Simu_Liu_image = face_recognition.load_image_file('images/New Sample/Simu_Liu.jpg')
    Simu_Liu_encodings = face_recognition.face_encodings(Simu_Liu_image)[0]

    # Face #18
    taylor_swift_image = face_recognition.load_image_file('images/New Sample/taylor-swift.jpg')
    taylor_swift_encodings = face_recognition.face_encodings(taylor_swift_image)[0]

    # Face #19
    the_rock_image = face_recognition.load_image_file('images/New Sample/the-rock.jpg')
    the_rock_encodings = face_recognition.face_encodings(the_rock_image)[0]

    # Face #20
    Tom_Holland_image = face_recognition.load_image_file('images/New Sample/Tom_Holland.jpg')
    Tom_Holland_encodings = face_recognition.face_encodings(Tom_Holland_image)[0]

    # Face #21
    modi_image = face_recognition.load_image_file('images/modi.jpg')
    modi_encodings = face_recognition.face_encodings(modi_image)[0]

    # #save the encodings and the corresponding labels in seperate arrays in the same order
    known_face_encodings = [Benedict_Cumberbatch_face_encodings, Beyonce_face_encodings, lim_face_encodings,
                            Brie_Larson_encodings, Calum_Scott_encodings, Chris_Hemsworth_encodings,
                            Ed_Sheeran_encodings, Evangeline_Lilly_encodings,
                            Hailey_Rhode_Bieber_encodings, justin_bieber_encodings, lisa_encodings,
                            Mahershala_Ali_encodings,
                            Paul_Rudd_encodings, Ryan_Reynolds_encodings, Scarlett_Johansson_encodings,
                            Simu_Liu_encodings, taylor_swift_encodings,
                            the_rock_encodings, Tom_Holland_encodings, modi_encodings]
    known_face_names = ["Benedict Cumberbatch", "Beyonce", "Lim Kah Yee",
                        "Brie Larson", "Calum Scott", "Chris Hemsworth", "Ed Sheeran", "Evangeline Lilly",
                        "Hailey Rhode Bieber", "Justin Bieber", "Lisa", " Mahershala Ali", "Paul Rudd",
                        "Ryan Reynolds", "Scarlett Johansson", "Simu Liu", "Taylor Swift", "Johnson Dwayne",
                        "Tom Holland", "Narendra Modi"]

    # initialize the array variable to hold all face locations, encodings and names
    all_face_locations = []
    all_face_encodings = []
    all_face_names = []

    # loop through every frame in the video
    while file_video_stream.isOpened:
        # get the current frame from the video stream as an image
        ret, current_frame = file_video_stream.read()

        if ret:
            #Gender Classification
            gray = cv2.cvtColor(current_frame, cv2.COLOR_RGB2GRAY)

            # step-3: crop the face (using haar cascase classifier)
            faces = haar.detectMultiScale(gray, 1.5, 3)
            
            # resize the current frame to 1/4 size to proces faster
            current_frame_small = cv2.resize(current_frame, (0, 0), fx=0.25, fy=0.25)
            # detect all faces in the image
            # arguments are image,no_of_times_to_upsample, model
            all_face_locations = face_recognition.face_locations(current_frame_small, number_of_times_to_upsample=1,
                                                                 model='hog')

            # detect face encodings for all the faces detected
            all_face_encodings = face_recognition.face_encodings(current_frame_small, all_face_locations)

            # looping through the face locations and the face embeddings
            for current_face_location, current_face_encoding in zip(all_face_locations, all_face_encodings):
                # splitting the tuple to get the four position values of current face
                top_pos, right_pos, bottom_pos, left_pos = current_face_location

                for x, y, w, h in faces:
                    roi = gray[y:y + h, x:x + w]  # crop image
                    # step - 4: normalization (0-1)
                    roi = roi / 255.0
                    # step-5: resize images (100,100)
                    if roi.shape[1] > 100:
                        roi_resize = cv2.resize(roi, (100, 100), cv2.INTER_AREA)
                    else:
                        roi_resize = cv2.resize(roi, (100, 100), cv2.INTER_CUBIC)
                    # step-6: Flattening (1x10000)
                    roi_reshape = roi_resize.reshape(1, 10000)  # 1,-1
                    # step-7: subptract with mean
                    roi_mean = roi_reshape - mean
                    # step -8: get eigen image
                    eigen_image = model_pca.transform(roi_mean)
                    # step -9: pass to ml model (svm)
                    results = model_svm.predict_proba(eigen_image)[0]
                    # step -10:
                    pre = results.argmax()  # 0 or 1
                    score = results[pre]

                    textPreGender = "%s : %0.2f" % (gender_pre[pre], score)
                    cv2.putText(current_frame, textPreGender, (x, y), cv2.FONT_HERSHEY_DUPLEX, 1, (255, 255, 0), 2)
                # change the position maginitude to fit the actual size video frame
                top_pos = top_pos * 4
                right_pos = right_pos * 4
                bottom_pos = bottom_pos * 4
                left_pos = left_pos * 4

                # find all the matches and get the list of matches
                all_matches = face_recognition.compare_faces(known_face_encodings, current_face_encoding)

                # string to hold the label
                name_of_person = 'Unknown face'

                # check if the all_matches have at least one item
                # if yes, get the index number of face that is located in the first index of all_matches
                # get the name corresponding to the index number and save it in name_of_person
                if True in all_matches:
                    first_match_index = all_matches.index(True)
                    name_of_person = known_face_names[first_match_index]
                # draw rectangle around the face
                cv2.rectangle(current_frame, (left_pos, top_pos), (right_pos, bottom_pos), (0, 255, 0), 2)

                # display the name as text in the image
                font = cv2.FONT_HERSHEY_DUPLEX
                
                cv2.putText(current_frame, name_of_person, (left_pos, bottom_pos), font, 1, (255, 255, 0), 2, cv2.LINE_AA)

            # display the video
            cv2.imshow("Webcam Video", current_frame)

            if cv2.waitKey(1) & 0xFF == ord('q'):
                break
        else:
            break

    # release the stream and cam
    # close all opencv windows open
    file_video_stream.release()
    cv2.destroyAllWindows()



def faceRecognitionReal():
    # capture the video from default camera
    webcam_video_stream = cv2.VideoCapture(0)

    # Face #1
    Benedict_Cumberbatch_image = face_recognition.load_image_file('images/New Sample/Benedict_Cumberbatch.jpg')
    Benedict_Cumberbatch_face_encodings = face_recognition.face_encodings(Benedict_Cumberbatch_image)[0]

    # Face #2
    Beyonce_image = face_recognition.load_image_file('images/New Sample/Beyonce.jpg')
    Beyonce_face_encodings = face_recognition.face_encodings(Beyonce_image)[0]

    # Face #3
    lim_image = face_recognition.load_image_file('images/Lim.jpeg')
    lim_face_encodings = face_recognition.face_encodings(lim_image)[0]

    # Face #4
    Brie_Larson_image = face_recognition.load_image_file('images/New Sample/Brie_Larson.jpg')
    Brie_Larson_encodings = face_recognition.face_encodings(Brie_Larson_image)[0]

    # Face #5
    Calum_Scott_image = face_recognition.load_image_file('images/New Sample/Calum  Scott.jpeg')
    Calum_Scott_encodings = face_recognition.face_encodings(Calum_Scott_image)[0]

    # Face #6
    Chris_Hemsworth_image = face_recognition.load_image_file('images/New Sample/Chris Hemsworth.jpg')
    Chris_Hemsworth_encodings = face_recognition.face_encodings(Chris_Hemsworth_image)[0]

    # Face #7
    Chris_Hemsworth_image = face_recognition.load_image_file('images/New Sample/Chris Pratt.jpg')
    Chris_Hemsworth_encodings = face_recognition.face_encodings(Chris_Hemsworth_image)[0]

    # Face #8
    Ed_Sheeran_image = face_recognition.load_image_file('images/New Sample/Ed Sheeran.png')
    Ed_Sheeran_encodings = face_recognition.face_encodings(Ed_Sheeran_image)[0]

    # Face #9
    Evangeline_Lilly_image = face_recognition.load_image_file('images/New Sample/Evangeline_Lilly.jpg')
    Evangeline_Lilly_encodings = face_recognition.face_encodings(Evangeline_Lilly_image)[0]

    # Face #10
    Hailey_Rhode_Bieber_image = face_recognition.load_image_file('images/New Sample/Hailey Rhode Bieber.jpg')
    Hailey_Rhode_Bieber_encodings = face_recognition.face_encodings(Hailey_Rhode_Bieber_image)[0]

    # Face #11
    justin_bieber_image = face_recognition.load_image_file('images/New Sample/justin bieber.jpg')
    justin_bieber_encodings = face_recognition.face_encodings(justin_bieber_image)[0]

    # Face #12
    lisa_image = face_recognition.load_image_file('images/New Sample/lisa.png')
    lisa_encodings = face_recognition.face_encodings(lisa_image)[0]

    # Face #13
    Mahershala_Ali_image = face_recognition.load_image_file('images/New Sample/Mahershala_Ali.jpg')
    Mahershala_Ali_encodings = face_recognition.face_encodings(Mahershala_Ali_image)[0]

    # Face #14
    Paul_Rudd_image = face_recognition.load_image_file('images/New Sample/Paul_Rudd.jpg')
    Paul_Rudd_encodings = face_recognition.face_encodings(Paul_Rudd_image)[0]

    # Face #14
    Paul_Rudd_image = face_recognition.load_image_file('images/New Sample/Richard_Madden.jpg')
    Paul_Rudd_encodings = face_recognition.face_encodings(Paul_Rudd_image)[0]

    # Face #15
    Ryan_Reynolds_image = face_recognition.load_image_file('images/New Sample/Ryan_Reynolds.jpg')
    Ryan_Reynolds_encodings = face_recognition.face_encodings(Ryan_Reynolds_image)[0]

    # Face #16
    Scarlett_Johansson_image = face_recognition.load_image_file('images/New Sample/Scarlett Johansson.jpg')
    Scarlett_Johansson_encodings = face_recognition.face_encodings(Scarlett_Johansson_image)[0]

    # Face #17
    Simu_Liu_image = face_recognition.load_image_file('images/New Sample/Simu_Liu.jpg')
    Simu_Liu_encodings = face_recognition.face_encodings(Simu_Liu_image)[0]

    # Face #18
    taylor_swift_image = face_recognition.load_image_file('images/New Sample/taylor-swift.jpg')
    taylor_swift_encodings = face_recognition.face_encodings(taylor_swift_image)[0]

    # Face #19
    the_rock_image = face_recognition.load_image_file('images/New Sample/the-rock.jpg')
    the_rock_encodings = face_recognition.face_encodings(the_rock_image)[0]

    # Face #20
    Tom_Holland_image = face_recognition.load_image_file('images/New Sample/Tom_Holland.jpg')
    Tom_Holland_encodings = face_recognition.face_encodings(Tom_Holland_image)[0]

    # Face #21
    modi_image = face_recognition.load_image_file('images/modi.jpg')
    modi_encodings = face_recognition.face_encodings(modi_image)[0]

    # #save the encodings and the corresponding labels in seperate arrays in the same order
    known_face_encodings = [Benedict_Cumberbatch_face_encodings, Beyonce_face_encodings, lim_face_encodings,
                            Brie_Larson_encodings, Calum_Scott_encodings, Chris_Hemsworth_encodings,
                            Ed_Sheeran_encodings, Evangeline_Lilly_encodings,
                            Hailey_Rhode_Bieber_encodings, justin_bieber_encodings, lisa_encodings,
                            Mahershala_Ali_encodings,
                            Paul_Rudd_encodings, Ryan_Reynolds_encodings, Scarlett_Johansson_encodings,
                            Simu_Liu_encodings, taylor_swift_encodings,
                            the_rock_encodings, Tom_Holland_encodings, modi_encodings]
    known_face_names = ["Benedict Cumberbatch", "Beyonce", "Lim Kah Yee",
                        "Brie Larson", "Calum Scott", "Chris Hemsworth", "Ed Sheeran", "Evangeline Lilly",
                        "Hailey Rhode Bieber", "Justin Bieber", "Lisa", " Mahershala Ali", "Paul Rudd",
                        "Ryan Reynolds", "Scarlett Johansson", "Simu Liu", "Taylor Swift", "Johnson Dwayne",
                        "Tom Holland", "Narendra Modi"]

    # initialize the array variable to hold all face locations, encodings and names
    all_face_locations = []
    all_face_encodings = []
    all_face_names = []

    # loop through every frame in the video
    while True:
        # get the current frame from the video stream as an image
        ret, current_frame = webcam_video_stream.read()

        #Gender Classification
        gray = cv2.cvtColor(current_frame, cv2.COLOR_RGB2GRAY)

        # step-3: crop the face (using haar cascase classifier)
        faces = haar.detectMultiScale(gray, 1.5, 3)

        # resize the current frame to 1/4 size to proces faster
        current_frame_small = cv2.resize(current_frame, (0, 0), fx=0.25, fy=0.25)
        # detect all faces in the image
        # arguments are image,no_of_times_to_upsample, model
        all_face_locations = face_recognition.face_locations(current_frame_small, number_of_times_to_upsample=1,
                                                             model='hog')

        # detect face encodings for all the faces detected
        all_face_encodings = face_recognition.face_encodings(current_frame_small, all_face_locations)

        # looping through the face locations and the face embeddings
        for current_face_location, current_face_encoding in zip(all_face_locations, all_face_encodings):

            for x, y, w, h in faces:
                roi = gray[y:y + h, x:x + w]  # crop image
                # step - 4: normalization (0-1)
                roi = roi / 255.0
                # step-5: resize images (100,100)
                if roi.shape[1] > 100:
                    roi_resize = cv2.resize(roi, (100, 100), cv2.INTER_AREA)
                else:
                    roi_resize = cv2.resize(roi, (100, 100), cv2.INTER_CUBIC)
                # step-6: Flattening (1x10000)
                roi_reshape = roi_resize.reshape(1, 10000)  # 1,-1
                # step-7: subptract with mean
                roi_mean = roi_reshape - mean
                # step -8: get eigen image
                eigen_image = model_pca.transform(roi_mean)
                # step -9: pass to ml model (svm)
                results = model_svm.predict_proba(eigen_image)[0]
                # step -10:
                pre = results.argmax()  # 0 or 1
                score = results[pre]

                textPreGender = "%s : %0.2f" % (gender_pre[pre], score)
                cv2.putText(current_frame, textPreGender, (x, y), cv2.FONT_HERSHEY_DUPLEX, 1, (255, 255, 0), 2)
            # splitting the tuple to get the four position values of current face
            top_pos, right_pos, bottom_pos, left_pos = current_face_location

            # change the position maginitude to fit the actual size video frame
            top_pos = top_pos * 4
            right_pos = right_pos * 4
            bottom_pos = bottom_pos * 4
            left_pos = left_pos * 4

            # find all the matches and get the list of matches
            all_matches = face_recognition.compare_faces(known_face_encodings, current_face_encoding)

            # string to hold the label
            name_of_person = 'Unknown face'

            # check if the all_matches have at least one item
            # if yes, get the index number of face that is located in the first index of all_matches
            # get the name corresponding to the index number and save it in name_of_person
            if True in all_matches:
                first_match_index = all_matches.index(True)
                name_of_person = known_face_names[first_match_index]

            # draw rectangle around the face
            cv2.rectangle(current_frame, (left_pos, top_pos), (right_pos, bottom_pos), (255, 0, 0), 2)

            # display the name as text in the image
            font = cv2.FONT_HERSHEY_DUPLEX
            cv2.putText(current_frame, name_of_person, (left_pos, bottom_pos), font, 0.5, (255, 255, 255), 1)

        img = cv2.imencode('.jpg', current_frame)[1].tobytes()
        yield b'--frame\r\n'b'Content-Type: image/jpeg\r\n\r\n' + img + b'\r\n'


# Object Detection
def mobileNetImageDetection(path, filename):
    # use for return the image name
    imgName = filename.split('.')[0]

    thres = 0.45  # Threshold to detect object

    classNames = []
    classFile = 'mobileNet/coco.names'
    with open(classFile, 'rt') as f:
        classNames = f.read().rstrip('\n').split('\n')

    net = initNet()

    img = cv2.imread(path)
    classIds, confs, bbox = net.detect(img, confThreshold=thres)

    if len(classIds) != 0:
        for classId, confidence, box in zip(classIds.flatten(), confs.flatten(), bbox):

            if confidence >= 0.5:

                cv2.rectangle(img, box, color=(0, 255, 0), thickness=2)

                if (box[2] - box[0]) < 400:
                    cv2.putText(img, classNames[classId - 1].upper(), (box[0] + 10, box[1] + 30),
                                cv2.FONT_HERSHEY_COMPLEX, 0.5, (0, 255, 0), 1)

                    cv2.putText(img, str(round(confidence * 100, 2)), (box[0] + 120, box[1] + 30),
                                cv2.FONT_HERSHEY_COMPLEX, 0.5, (0, 255, 0), 1)

                else:
                    cv2.putText(img, classNames[classId - 1].upper(), (box[0] + 10, box[1] + 30),
                                cv2.FONT_HERSHEY_COMPLEX, 1, (0, 255, 0), 1)

                    cv2.putText(img, str(round(confidence * 100, 2)), (box[0] + 200, box[1] + 30),
                                cv2.FONT_HERSHEY_COMPLEX, 1, (0, 255, 0), 1)

                imgName = imgName + '_' + classNames[classId - 1].capitalize() + '_' + str(
                    round(confidence * 100, 2)) + '%'

                exportImage = os.path.join(
                    os.path.join(pathlib.Path().parent.absolute(), OBJECT_IMAGE_EXPORT_PATH, imgName + '.jpg'))
                cv2.imwrite(exportImage, img)

    return imgName


def mobileNetVideoDetection(path):
    thres = 0.45
    cap = cv2.VideoCapture(path)

    cap.set(3, 1280)
    cap.set(4, 720)
    cap.set(10, 70)

    classNames = []
    classFile = 'mobileNet/coco.names'
    with open(classFile, 'rt') as f:
        classNames = f.read().rstrip('\n').split('\n')

    net = initNet()

    createWorkBook()

    wb = load_workbook('ObjectDetected.xlsx')
    sheet = wb.worksheets[0]
    print("Clearing Data")
    while sheet.max_row > 1:
        # this method removes the row 2
        sheet.delete_rows(2)
        wb.save('ObjectDetected.xlsx')
    print("Data Cleared")

    # fps
    prevTime = 0

    while cap.isOpened:
        success, img = cap.read()

        if success:
            img = cv2.resize(img, (800, 600))

            classIds, confs, bbox = net.detect(img, confThreshold=thres)

            if len(classIds) != 0:
                for classId, confidence, box in zip(classIds.flatten(), confs.flatten(), bbox):

                    if confidence >= 0.5:
                        cv2.rectangle(img, box, color=(0, 255, 0), thickness=2)

                        if (box[2] - box[0]) < 400:
                            cv2.putText(img, classNames[classId - 1].upper(), (box[0] + 10, box[1] + 30),
                                        cv2.FONT_HERSHEY_COMPLEX, 0.5, (0, 255, 0), 1)

                            cv2.putText(img, str(round(confidence * 100, 2)), (box[0] + 120, box[1] + 30),
                                        cv2.FONT_HERSHEY_COMPLEX, 0.5, (0, 255, 0), 1)
                        else:
                            cv2.putText(img, classNames[classId - 1].upper(), (box[0] + 10, box[1] + 30),
                                        cv2.FONT_HERSHEY_COMPLEX, 1, (0, 255, 0), 1)

                            cv2.putText(img, str(round(confidence * 100, 2)), (box[0] + 200, box[1] + 30),
                                        cv2.FONT_HERSHEY_COMPLEX, 1, (0, 255, 0), 1)

                        # Append detected object into excel file
                        objectDetected = (classNames[classId - 1].capitalize(), round(confidence * 100, 2))
                        sheet.append(objectDetected)
                        wb.save('ObjectDetected.xlsx')

                currentTime = time.time()
                fps = 1 / (currentTime - prevTime)
                prevTime = currentTime

                cv2.putText(img, "FPS: " + str(round(fps, 2)), (10, 50), font, 1, (0, 0, 0), 3)

            cv2.imshow("Output", img)
            if cv2.waitKey(1) & 0xFF == ord('s'):  # press s to exit  --#esc key (27),
                break
        else:
            break

    cap.release()
    cv2.destroyAllWindows()


# Finger Images
folderPath = "FingerImages"
myList = os.listdir(folderPath)
overlayList = []

# create mediapipe
detector = htm.handDetector(detectionCon=0.75)

# hand index
tipIds = [4, 8, 12, 16, 20]


def retrieveFingerImage(path, fingerList):
    for imgPath in fingerList:
        image = cv2.imread(f'{path}/{imgPath}')

        overlayList.append(image)

    return overlayList


def mobileRealTimeDetection():
    overlayList = retrieveFingerImage(folderPath, myList)

    thres = 0.45  # Threshold to detect object

    # Start Real Time
    cap = cv2.VideoCapture(0)

    classNames = []
    classFile = 'mobileNet/coco.names'
    with open(classFile, 'rt') as f:
        classNames = f.read().rstrip('\n').split('\n')

    net = initNet()

    # fps
    font = cv2.FONT_HERSHEY_PLAIN
    starting_time = time.time()
    frame_id = 0

    # Close due to low fps
    # createWorkBook()
    #
    # wb = load_workbook('ObjectDetected.xlsx')
    # sheet = wb.worksheets[0]
    # print("Clearing Data")
    # while sheet.max_row > 1:
    #     # this method removes the row 2
    #     sheet.delete_rows(2)
    #     wb.save('ObjectDetected.xlsx')
    # print("Data Cleared")

    while True:
        success, img = cap.read()
        img = cv2.resize(img, (0, 0), fx=1.05, fy=1.0)

        frame_id += 1

        classIds, confs, bbox = net.detect(img, confThreshold=thres)

        if len(classIds) != 0:
            for classId, confidence, box in zip(classIds.flatten(), confs.flatten(), bbox):
                if confidence >= 0.5:
                    cv2.rectangle(img, box, color=(0, 255, 0), thickness=2)

                    cv2.putText(img, classNames[classId - 1].upper(), (box[0] + 10, box[1] + 30),
                                cv2.FONT_HERSHEY_COMPLEX, 1, (0, 255, 0), 1)

                    cv2.putText(img, str(round(confidence * 100, 2)), (box[0] + 200, box[1] + 30),
                                cv2.FONT_HERSHEY_COMPLEX, 1, (0, 255, 0), 1)

                    if classNames[classId - 1].upper() == "PERSON":
                        img = detector.findHands(img)
                        lmList = detector.findPosition(img, draw=False)

                        if len(lmList) != 0:
                            fingers = []

                            # thumb
                            if lmList[5][1] > lmList[17][1]:
                                #  Right
                                # [index finger][height]
                                if lmList[tipIds[0]][1] > lmList[tipIds[0] - 1][2] + 100:
                                    fingers.append(1)
                                else:
                                    fingers.append(0)
                            else:
                                # Left
                                # [index finger][height]
                                if lmList[tipIds[0]][1] < lmList[tipIds[0] - 1][2]:
                                    fingers.append(1)
                                else:
                                    fingers.append(0)

                            # 4 fingers
                            for id in range(1, 5):
                                # [index finger][height]
                                if lmList[tipIds[id]][2] < lmList[tipIds[id] - 2][2]:
                                    fingers.append(1)
                                else:
                                    fingers.append(0)

                            totalFingers = fingers.count(1)

                            h, w, c = overlayList[totalFingers - 1].shape
                            img[0:h, 0:w] = overlayList[totalFingers - 1]

                    # Close due to low fps
                    # Append detected object into excel file
                    # objectDetected = (classNames[classId - 1].capitalize(), round(confidence * 100, 2))
                    # sheet.append(objectDetected)
                    # wb.save('ObjectDetected.xlsx')

                elapsed_time = time.time() - starting_time
                fps = frame_id / elapsed_time
                cv2.putText(img, "FPS: " + str(round(fps, 2)), (10, 50), font, 2, (0, 0, 0), 3)

                frame = cv2.imencode('.jpg', img)[1].tobytes()
                yield b'--frame\r\n'b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n'


def initNet():
    configPath = 'mobileNet/ssd_mobilenet_v3_large_coco_2020_01_14.pbtxt'
    weightsPath = 'mobileNet/frozen_inference_graph.pb'

    net = cv2.dnn_DetectionModel(weightsPath, configPath)
    net.setInputSize(320, 320)
    net.setInputScale(1.0 / 127.5)
    net.setInputMean((127.5, 127.5, 127.5))
    net.setInputSwapRB(True)

    return net


# Need to manually create by add method to the code
def createWorkBook():
    if not os.path.isfile('ObjectDetected.xlsx'):
        objectDetectedBook = xlsxwriter.Workbook('ObjectDetected.xlsx')
        objectSheet = objectDetectedBook.add_worksheet()
        objectSheet.write(0, 0, 'Object Detected')
        objectSheet.write(0, 1, 'Percentage')

        objectDetectedBook.close()
